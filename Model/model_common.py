import os
import sys
import datetime
import openpyxl
import tempfile
import win32api
from queue import Queue
from Model import (_ThreadModel, _DataBaseManager)
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


class ErrorReportToDB(_ThreadModel):
    def __init__(self, _thread_queue: Queue, _manager_index: int, _error_info: dict, _app_queue: [Queue, None] = None):
        self._code = 'ModelCommon_ErrorReportToDB'
        self._manager_index, self._error_info, self._dbm = _manager_index, _error_info, _DataBaseManager(self._code)
        super().__init__(_thread_queue, _app_queue)

    def run(self):
        try:
            _dcr = self._dbm.connecting()
            if _dcr[0]:
                _query = '''
                INSERT INTO error_report(manager_index, error_code, error_class, error_description, error_line, error_data)
                VALUES({}, '{}', '{}', '{}', {}, '{}')
                '''.format(
                    self._manager_index,
                    self._error_info['ECD'],
                    str(self._error_info['CLS']).replace('\"', '\'').replace('\'', ''),
                    str(self._error_info['DES']).replace('\"', '\'').replace('\'', ''),
                    self._error_info['LNO'],
                    str(self._error_info['DTA']).replace('\"', '\'').replace('\'', '')
                )
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    self.send(False, True)
                else:
                    self.send(False, False, _der[2])
            else:
                self.send(False, False, _dcr[1])
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': []
            })
        finally:
            self._dbm.disconnecting()


class ErrorReportToLocal(_ThreadModel):
    def __init__(self, _thread_queue: Queue, _manager_index: int, _origin_error_info: dict, _local_master_path: str, _second_error_info: dict = None, _app_queue: [Queue, None] = None):
        self._code = 'ModelCommon_ErrorReportToLocal'
        self._manager_index, self._origin_error_info, self._second_error_info, self._local_master_path = _manager_index, _origin_error_info, _second_error_info, _local_master_path
        super().__init__(_thread_queue, _app_queue)

    def run(self):
        _local_file_path = '{}/{}.txt'.format(self._local_master_path, datetime.datetime.strftime(datetime.datetime.now(), "%Y%m%d%H%M%S"))
        try:
            with open(_local_file_path, 'w') as _local_file:
                _local_file.write('ERROR OCCURRED USER : ' + str(self._manager_index) + '\n')
                _local_file.write('ERROR OCCURRED TIME : ' + str(datetime.datetime.now()) + '\n')
                _local_file.write('ERROR OCCURRED CODE : ' + self._origin_error_info['ECD'] + '\n')
                _local_file.write('ERROR OCCURRED CLSS : ' + str(self._origin_error_info['CLS']) + '\n')
                _local_file.write('ERROR OCCURRED DESC : ' + str(self._origin_error_info['DES']) + '\n')
                _local_file.write('ERROR OCCURRED LNUM : ' + str(self._origin_error_info['LNO']) + '\n')
                _local_file.write('ERROR OCCURRED DATA : ' + str(self._origin_error_info['DTA']) + '\n')
                _local_file.write('====================================================================================================\n')
                _local_file.write('ERROR OCCURRED USER : ' + str(self._manager_index) + '\n')
                _local_file.write('ERROR OCCURRED TIME : ' + str(datetime.datetime.now()) + '\n')
                _local_file.write('ERROR OCCURRED CODE : ' + self._second_error_info['ECD'] + '\n')
                _local_file.write('ERROR OCCURRED CLSS : ' + str(self._second_error_info['CLS']) + '\n')
                _local_file.write('ERROR OCCURRED DESC : ' + str(self._second_error_info['DES']) + '\n')
                _local_file.write('ERROR OCCURRED LNUM : ' + str(self._second_error_info['LNO']) + '\n')
                _local_file.write('ERROR OCCURRED DATA : ' + str(self._second_error_info['DTA']) + '\n')
                _local_file.write('====================================================================================================\n')
            self.send(False, True, _local_file_path)
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [_local_file_path]
            })


class GetLookUpDataDetail(_ThreadModel):
    def __init__(self, _thread_queue: Queue, _manager_index: int, _info_core_index: int, _app_queue: [Queue, None] = None):
        self._code, self._manager_index, self._info_core_index = 'ModelCommon_GetLookUpDataDetail', _manager_index, _info_core_index
        self._dbm = _DataBaseManager(self._code)
        super().__init__(_thread_queue, _app_queue)

    def run(self):
        try:
            _data = {}
            _dcr = self._dbm.connecting()
            if _dcr[0]:
                _query = '''
                SELECT DATE_FORMAT(LOOKUP_DATE, '%Y-%m-%d'),
                       FORMAT_EXPORT_NUMBER,
                       CHASSIS_NUMBER,
                       DATE_FORMAT(REPORT_DATE, '%Y-%m-%d'),
                       DATE_FORMAT(LOAD_DUTY_DEADLINE, '%Y-%m-%d'),
                       VEHICLE_STATUS,
                       TOTAL_AMOUNT,
                       TOTAL_WEIGHT,
                       PRE_AMOUNT,
                       PRE_WEIGHT,
                       REMAIN_AMOUNT,
                       REMAIN_WEIGHT,
                       FLAG_INSPECTION,
                       EXPORT_SHIPPER,
                       MAKER,
                       DATE_FORMAT(ACCEPT_DATE, '%Y-%m-%d'),
                       FLAG_PRE_SHIPPING,
                       SHIP_NAME
                FROM merged_ei_cn_en
                WHERE KEY_INDEX = {};
                '''.format(self._info_core_index)
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    _cur_data = _der[1][0]
                else:
                    self.send(False, False, _der[2])
                    return
                _query = '''
                SELECT bl_number,
                       DATE_FORMAT(shipment_date, '%Y-%m-%d'),
                       pre_shipping_amount,
                       pre_shipping_weight
                FROM export_info_export_number_bl
                         JOIN dict_bl_number dbn on export_info_export_number_bl.d_bl_index = dbn.d_bl_index
                WHERE ei_export_index IN (SELECT ei_export_index FROM lookup_chassis_result WHERE r_chassis_index = {});
                '''.format(self._info_core_index)
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    _cur_bl_data = _der[1]
                else:
                    self.send(False, False, _der[2])
                    return
                _data.__setitem__(
                    'CUR_DATA',
                    [
                        {
                            '조회일자': _cur_data[0]
                        },
                        {
                            '수출신고번호': _cur_data[1],
                            '차대번호': _cur_data[2],
                            '신고일자': _cur_data[3],
                            '적재의무기한': _cur_data[4],
                            '차량진행상태': _cur_data[5]
                        },
                        {
                            '총 개수': _cur_data[6],
                            '총 중량': _cur_data[7]
                        },
                        {
                            '선적 개수': _cur_data[8],
                            '선적 중량': _cur_data[9]
                        },
                        {
                            '잔여 개수': _cur_data[10],
                            '잔여 중량': _cur_data[11]
                        },
                        {
                            '검사유무': _cur_data[12]
                        },
                        {
                            '수출화주/대행자': _cur_data[13],
                            '제조자': _cur_data[14],
                            '수리일자': _cur_data[15],
                            '선기적완료여부': _cur_data[16],
                            '선박/편명': _cur_data[17]
                        },
                        [list(_row) for _row in _cur_bl_data]
                    ]
                )
                _query = '''SELECT GET_COMPARE_LOOKUP_CHASSIS_RESULT({}, {});'''.format(self._info_core_index, self._manager_index)
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    _status_flag = int(_der[1][0][0])
                else:
                    self.send(False, False, _der[2])
                    return
                _data.__setitem__('STATUS_FLAG', _status_flag)
                if _status_flag == 2:
                    _query = '''
                    SELECT KEY_INDEX
                    FROM merged_ei_cn_en
                    WHERE D_CHASSIS_INDEX = (SELECT D_CHASSIS_INDEX FROM merged_ei_cn_en WHERE KEY_INDEX = {0})
                      AND LOOKUP_CHASSIS_INDEX < (SELECT LOOKUP_CHASSIS_INDEX FROM merged_ei_cn_en WHERE KEY_INDEX = {0})
                    ORDER BY KEY_INDEX DESC LIMIT 1;
                    '''.format(self._info_core_index)
                    _der = self._dbm.execute_query(_query)
                    if _der[0]:
                        _com_info_core_index = _der[1][0][0]
                    else:
                        self.send(False, False, _der[2])
                        return
                    _query = '''
                    SELECT DATE_FORMAT(LOOKUP_DATE, '%Y-%m-%d'),
                           FORMAT_EXPORT_NUMBER,
                           CHASSIS_NUMBER,
                           DATE_FORMAT(REPORT_DATE, '%Y-%m-%d'),
                           DATE_FORMAT(LOAD_DUTY_DEADLINE, '%Y-%m-%d'),
                           VEHICLE_STATUS,
                           TOTAL_AMOUNT,
                           TOTAL_WEIGHT,
                           PRE_AMOUNT,
                           PRE_WEIGHT,
                           REMAIN_AMOUNT,
                           REMAIN_WEIGHT,
                           FLAG_INSPECTION,
                           EXPORT_SHIPPER,
                           MAKER,
                           DATE_FORMAT(ACCEPT_DATE, '%Y-%m-%d'),
                           FLAG_PRE_SHIPPING,
                           SHIP_NAME
                    FROM merged_ei_cn_en
                    WHERE KEY_INDEX = {};
                    '''.format(_com_info_core_index)
                    _der = self._dbm.execute_query(_query)
                    if _der[0]:
                        _com_data = _der[1][0]
                    else:
                        self.send(False, False, _der[2])
                        return
                    _query = '''
                    SELECT bl_number,
                           DATE_FORMAT(shipment_date, '%Y-%m-%d'),
                           pre_shipping_amount,
                           pre_shipping_weight
                    FROM export_info_export_number_bl
                             JOIN dict_bl_number dbn on export_info_export_number_bl.d_bl_index = dbn.d_bl_index
                    WHERE ei_export_index IN (SELECT ei_export_index FROM lookup_chassis_result WHERE r_chassis_index = {});
                    '''.format(_com_info_core_index)
                    _der = self._dbm.execute_query(_query)
                    if _der[0]:
                        _com_bl_data = _der[1]
                    else:
                        self.send(False, False, _der[2])
                        return
                    _data.__setitem__(
                        'COM_DATA',
                        [
                            {
                                '조회일자': _com_data[0]
                            },
                            {
                                '수출신고번호': _com_data[1],
                                '차대번호': _com_data[2],
                                '신고일자': _com_data[3],
                                '적재의무기한': _com_data[4],
                                '차량진행상태': _com_data[5]
                            },
                            {
                                '총 개수': _com_data[6],
                                '총 중량': _com_data[7]
                            },
                            {
                                '선적 개수': _com_data[8],
                                '선적 중량': _com_data[9]
                            },
                            {
                                '잔여 개수': _com_data[10],
                                '잔여 중량': _com_data[11]
                            },
                            {
                                '검사유무': _com_data[12]
                            },
                            {
                                '수출화주/대행자': _com_data[13],
                                '제조자': _com_data[14],
                                '수리일자': _com_data[15],
                                '선기적완료여부': _com_data[16],
                                '선박/편명': _com_data[17]
                            },
                            [list(_row) for _row in _com_bl_data]
                        ]
                    )
                else:
                    _data.__setitem__('COM_DATA', [])
                self.send(False, True, _data)
            else:
                self.send(False, False, _dcr[1])
                return
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [self._manager_index, self._info_core_index]
            })
        finally:
            self._dbm.disconnecting()


class DownloadVehicleDataToExcel(_ThreadModel):
    def __init__(
            self,
            _thread_queue: Queue,
            _download_path: str,
            _excel_name: str,
            _head: list,
            _inquiry_index: [int] = 0,
            _en_list: [list, None] = None,
            _cn_list: [list, None] = None,
            _vehicle_status: [str ,None] = None,
            _inspection_flag: [str, None] = None,
            _lookup_date_start: [str, None] = None,
            _lookup_date_close: [str, None] = None,
            _report_date_start: [str, None] = None,
            _report_date_close: [str ,None] = None,
            _load_duty_date_start: [str, None] = None,
            _load_duty_date_close: [str, None] = None,
            _exporter_shipper_name: [str, None] = None,
            _manager_index: [int, None] = None,
            _list_type: [int, None] = None,
            _app_queue: [Queue, None] = None
    ):
        self._code, self._download_path, self._excel_name, self._head = 'ModelCommon_DownloadVehicleDataToExcel', _download_path, _excel_name, _head + ['수출화주/대행자', '제조자', '수리일자', '선기적완료여부', '선박/편명']
        if _inquiry_index == 0:
            # Filter Data Query
            self._query_flag = 2
            self._en_list = _en_list
            self._cn_list = _cn_list
            self._vehicle_status = _vehicle_status
            self._inspection_flag = _inspection_flag
            self._lookup_date_start = _lookup_date_start
            self._lookup_date_close = _lookup_date_close
            self._report_date_start = _report_date_start
            self._report_date_close = _report_date_close
            self._load_duty_date_start = _load_duty_date_start
            self._load_duty_date_close = _load_duty_date_close
            self._exporter_shipper_name = _exporter_shipper_name
            self._manager_index = _manager_index
            self._list_type = _list_type
        else:
            # Inquiry Index Query
            self._query_flag = 1
            self._inquiry_index = _inquiry_index
        self._dbm = _DataBaseManager(self._code)
        self._excel: [openpyxl.Workbook, None] = None
        super().__init__(_thread_queue, _app_queue)

    def run(self):
        try:
            _dcr = self._dbm.connecting()
            if _dcr[0]:
                if self._query_flag == 1:
                    _query = '''
                    SELECT KEY_INDEX,
                           EXCEL_ROW_NO,
                           ORIGIN_EXPORT_NUMBER,
                           CHASSIS_NUMBER,
                           DATE_FORMAT(REPORT_DATE, '%Y-%m-%d'),
                           DATE_FORMAT(LOAD_DUTY_DEADLINE, '%Y-%m-%d'),
                           VEHICLE_STATUS,
                           TOTAL_AMOUNT,
                           TOTAL_WEIGHT,
                           PRE_AMOUNT,
                           PRE_WEIGHT,
                           REMAIN_AMOUNT,
                           REMAIN_WEIGHT,
                           FLAG_INSPECTION,
                           EXPORT_SHIPPER,
                           MAKER,
                           DATE_FORMAT(ACCEPT_DATE, '%Y-%m-%d'),
                           FLAG_PRE_SHIPPING,
                           SHIP_NAME
                    FROM merged_ei_cn_en
                    WHERE LOOKUP_CHASSIS_INDEX = {}
                    ORDER BY EXCEL_ROW_NO;
                    '''.format(self._inquiry_index)
                else:
                    _query = '''
                    SELECT KEY_INDEX,
                           0,
                           ORIGIN_EXPORT_NUMBER,
                           CHASSIS_NUMBER,
                           DATE_FORMAT(REPORT_DATE, '%Y-%m-%d'),
                           DATE_FORMAT(LOAD_DUTY_DEADLINE, '%Y-%m-%d'),
                           VEHICLE_STATUS,
                           TOTAL_AMOUNT,
                           TOTAL_WEIGHT,
                           PRE_AMOUNT,
                           PRE_WEIGHT,
                           REMAIN_AMOUNT,
                           REMAIN_WEIGHT,
                           FLAG_INSPECTION,
                           EXPORT_SHIPPER,
                           MAKER,
                           DATE_FORMAT(ACCEPT_DATE, '%Y-%m-%d'),
                           FLAG_PRE_SHIPPING,
                           SHIP_NAME
                    FROM merged_ei_cn_en
                    WHERE 1 = 1
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    '''.format(
                        'AND (KEY_INDEX) IN (SELECT MAX(KEY_INDEX) FROM merged_ei_cn_en GROUP BY CHASSIS_NUMBER, ORIGIN_EXPORT_NUMBER)' if self._list_type == 0 else '',
                        'AND (' + ' OR '.join(['ORIGIN_EXPORT_NUMBER LIKE \'%{}%\''.format(str(_).replace('-', '')) for _ in self._en_list]) + ')' if len(self._en_list) != 0 else '',
                        'AND (' + ' OR '.join(['CHASSIS_NUMBER LIKE \'%{}%\''.format(_) for _ in self._cn_list]) + ')' if len(self._cn_list) != 0 else '',
                        'AND VEHICLE_STATUS = \'{}\''.format(self._vehicle_status) if self._vehicle_status != 'All' else '',
                        'AND FLAG_INSPECTION = \'{}\''.format(self._inspection_flag) if self._inspection_flag != 'All' else '',
                        'AND (\'{}\' <= DATE_FORMAT(LOOKUP_DATE, \'%Y-%m-%d\') AND DATE_FORMAT(LOOKUP_DATE, \'%Y-%m-%d\') <= \'{}\')'.format(self._lookup_date_start, self._lookup_date_close),
                        'AND (\'{}\' <= REPORT_DATE AND REPORT_DATE <= \'{}\')'.format(self._report_date_start, self._report_date_close) if self._report_date_start != '' else '',
                        'AND (\'{}\' <= LOAD_DUTY_DEADLINE AND LOAD_DUTY_DEADLINE <= \'{}\')'.format(self._load_duty_date_start, self._load_duty_date_close) if self._load_duty_date_start != '' else '',
                        'AND EXPORT_SHIPPER LIKE \'%{}%\''.format(self._exporter_shipper_name) if self._exporter_shipper_name != '' else '',
                        'AND MANAGER_INDEX = {}'.format(self._manager_index) if self._manager_index != 0 else ''
                    )
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    self._excel = openpyxl.Workbook()
                    self._excel.active.title = 'Sheet'
                    _sheet = self._excel.active
                    for _col in range(1, len(self._head) + 1):
                        _sheet.cell(1, _col).value = self._head[_col - 1]
                    for _row in range(2, len(_der[1]) + 2):
                        _data_row = _der[1][_row - 2]
                        for _col in range(1, len(_data_row) + 1):
                            _sheet.cell(_row, _col).value = _data_row[_col - 1]
                    self._excel.save('{}/차량_{}'.format(self._download_path, self._excel_name))
                    self.send(False, True, '{}/차량_{}'.format(self._download_path, self._excel_name))
                else:
                    self.send(False, False, _der[2])
            else:
                self.send(False, False, _dcr[1])
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [self._inquiry_index] if self._query_flag == 1 else [self._en_list, self._cn_list, self._vehicle_status, self._inspection_flag, self._lookup_date_start, self._lookup_date_close, self._report_date_start, self._report_date_close, self._load_duty_date_start, self._load_duty_date_close, self._exporter_shipper_name, self._manager_index, self._list_type]
            })
        finally:
            if self._excel is not None:
                self._excel.close()
            self._dbm.disconnecting()


class DownloadBlDataToExcel(_ThreadModel):
    def __init__(
            self,
            _thread_queue: Queue,
            _download_path: str,
            _excel_name: str,
            _head: list,
            _lookup_index: [int] = 0,
            _bn_list: [list, None] = None,
            _en_list: [list, None] = None,
            _mn_list: [list, None] = None,
            _exporter_shipper_name: [str, None] = None,
            _lookup_date_start: [str, None] = None,
            _lookup_date_close: [str, None] = None,
            _accept_date_start: [str, None] = None,
            _accept_date_close: [str, None] = None,
            _load_duty_date_start: [str, None] = None,
            _load_duty_date_close: [str, None] = None,
            _departure_date_start: [str, None] = None,
            _departure_date_close: [str, None] = None,
            _shipment_place: [str, None] = None,
            _pre_shipping_flag: [str, None] = None,
            _manager_index: [int, None] = None,
            _list_type: [int, None] = None,
            _order_str: [str, None] = None,
            _app_queue: [Queue, None] = None
    ):
        self._code, self._download_path, self._excel_name, self._head = 'ModelCommon_DownloadBlDataToExcel', _download_path, _excel_name, _head
        if _lookup_index == 0:
            # Filter Data Query
            self._query_flag = 2
            self._bn_list = _bn_list
            self._en_list = _en_list
            self._mn_list = _mn_list
            self._exporter_shipper_name = _exporter_shipper_name
            self._lookup_date_start = _lookup_date_start
            self._lookup_date_close = _lookup_date_close
            self._accept_date_start = _accept_date_start
            self._accept_date_close = _accept_date_close
            self._load_duty_date_start = _load_duty_date_start
            self._load_duty_date_close = _load_duty_date_close
            self._departure_date_start = _departure_date_start
            self._departure_date_close = _departure_date_close
            self._shipment_place = _shipment_place
            self._pre_shipping_flag = _pre_shipping_flag
            self._manager_index = _manager_index
            self._list_type = _list_type
        else:
            self._query_flag = 1
            self._lookup_index = _lookup_index
        self._order_str = _order_str if _order_str != '' else 'ORDER BY KEY_INDEX DESC'
        self._dbm = _DataBaseManager(self._code)
        self._excel: [openpyxl.Workbook, None] = None
        super().__init__(_thread_queue, _app_queue)

    def run(self):
        try:
            _dcr = self._dbm.connecting()
            if _dcr[0]:
                if self._query_flag == 1:
                    _query = f'''
                    SELECT KEY_INDEX,
                           EXCEL_ROW_NO,
                           BL_NUMBER,
                           ORIGIN_EXPORT_NUMBER,
                           MANAGEMENT_NUMBER,
                           LOOKUP_DATE,
                           ACCEPT_DATE,
                           LOAD_DUTY_DEADLINE,
                           EXPORT_SHIPPER,
                           TOTAL_AMOUNT,
                           TOTAL_WEIGHT,
                           SHIPMENT_PLACE,
                           DEPARTURE_DATE,
                           PRE_AMOUNT,
                           PRE_WEIGHT,
                           DIVISION_COLLECT,
                           FLAG_PRE_SHIPPING
                    FROM merged_ei_bn
                    WHERE LOOKUP_BL_INDEX = {self._lookup_index}
                    {self._order_str};
                    '''
                else:
                    _query = '''
                    SELECT KEY_INDEX,
                           EXCEL_ROW_NO,
                           BL_NUMBER,
                           ORIGIN_EXPORT_NUMBER,
                           MANAGEMENT_NUMBER,
                           LOOKUP_DATE,
                           ACCEPT_DATE,
                           LOAD_DUTY_DEADLINE,
                           EXPORT_SHIPPER,
                           TOTAL_AMOUNT,
                           TOTAL_WEIGHT,
                           SHIPMENT_PLACE,
                           DEPARTURE_DATE,
                           PRE_AMOUNT,
                           PRE_WEIGHT,
                           DIVISION_COLLECT,
                           FLAG_PRE_SHIPPING
                    FROM merged_ei_bn
                    WHERE 1 = 1
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {}
                    {};
                    '''.format(
                        'AND (KEY_INDEX) IN (SELECT MAX(KEY_INDEX) FROM merged_ei_bn GROUP BY BL_NUMBER, ORIGIN_EXPORT_NUMBER)' if self._list_type == 0 else '',
                        'AND (' + ' OR '.join(['BL_NUMBER LIKE \'%{}%\''.format(_) for _ in self._bn_list]) + ')' if len(self._bn_list) != 0 else '',
                        'AND (' + ' OR '.join(['ORIGIN_EXPORT_NUMBER LIKE \'%{}%\''.format(str(_).replace('-', '')) for _ in self._en_list]) + ')' if len(self._en_list) != 0 else '',
                        'AND (' + ' OR '.join(['MANAGEMENT_NUMBER LIKE \'%{}%\''.format(_) for _ in self._mn_list]) + ')' if len(self._mn_list) != 0 else '',
                        'AND EXPORT_SHIPPER LIKE \'%{}%\''.format(self._exporter_shipper_name) if self._exporter_shipper_name != '' else '',
                        'AND (\'{}\' <= DATE_FORMAT(LOOKUP_DATE, \'%Y-%m-%d\') AND DATE_FORMAT(LOOKUP_DATE, \'%Y-%m-%d\') <= \'{}\')'.format(self._lookup_date_start, self._lookup_date_close),
                        'AND (\'{}\' <= ACCEPT_DATE AND ACCEPT_DATE <= \'{}\')'.format(self._accept_date_start, self._accept_date_close) if self._accept_date_start != '' else '',
                        'AND (\'{}\' <= LOAD_DUTY_DEADLINE AND LOAD_DUTY_DEADLINE <= \'{}\')'.format(self._load_duty_date_start, self._load_duty_date_close) if self._load_duty_date_start != '' else '',
                        'AND (\'{}\' <= DEPARTURE_DATE AND DEPARTURE_DATE <= \'{}\')'.format(self._departure_date_start, self._departure_date_close) if self._departure_date_start != '' else '',
                        'AND SHIPMENT_PLACE LIKE \'%{}%\''.format(self._shipment_place) if self._shipment_place != 'All' else '',
                        'AND FLAG_PRE_SHIPPING LIKE \'%{}%\''.format(self._pre_shipping_flag) if self._pre_shipping_flag != 'All' else '',
                        'AND MANAGER_INDEX = {}'.format(self._manager_index) if self._manager_index != 0 else '',
                        self._order_str
                    )
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    self._excel = openpyxl.Workbook()
                    self._excel.active.title = 'Sheet'
                    _sheet = self._excel.active
                    for _col in range(1, len(self._head) + 1):
                        _sheet.cell(1, _col).value = self._head[_col - 1]
                    for _row in range(2, len(_der[1]) + 2):
                        _data_row = _der[1][_row - 2]
                        for _col in range(1, len(_data_row) + 1):
                            _sheet.cell(_row, _col).value = _data_row[_col - 1]
                    self._excel.save('{}/BL_{}'.format(self._download_path, self._excel_name))
                    self.send(False, True, '{}/BL_{}'.format(self._download_path, self._excel_name))
                else:
                    self.send(False, False, _der[2])
            else:
                self.send(False, False, _dcr[1])
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [self._lookup_index] if self._query_flag == 1 else [self._bn_list, self._en_list, self._mn_list, self._exporter_shipper_name, self._lookup_date_start, self._lookup_date_close, self._accept_date_start, self._accept_date_close, self._load_duty_date_start, self._load_duty_date_close, self._departure_date_start, self._departure_date_close, self._shipment_place, self._pre_shipping_flag, self._manager_index, self._list_type]
            })
        finally:
            if self._excel is not None:
                self._excel.close()
            self._dbm.disconnecting()


class PrintBlDetail(_ThreadModel):
    def __init__(self, _thread_queue: Queue, _temp_path, _head_data, _body_data, _printer_name, _app_queue: [Queue, None] = None):
        self._code, self._temp_path, self._head_data, self._body_data, self._printer_name = 'ModelCommon_PrintBlDetail', _temp_path + '/' + 'temp_excel_file.xlsx', _head_data, _body_data, _printer_name
        super().__init__(_thread_queue, _app_queue)
        self._excel: [openpyxl.Workbook, None] = None
        self._sheet = None
        self._sheet_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        self._font_6 = Font(size=6)
        self._font_7 = Font(size=7)
        self._font_9 = Font(size=9)
        self._font_12 = Font(size=12)
        self._fill_gray = PatternFill(fgColor='F2F2F2', fill_type='solid')
        self._align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self._board = Border(
            left=Side(border_style='thin', color='00000000'),
            right=Side(border_style='thin', color='00000000'),
            top=Side(border_style='thin', color='00000000'),
            bottom=Side(border_style='thin', color='00000000')
        )

    def design_cell(self, _range: str, _value: [str, int], _font_size: Font, _fill: bool = False, _number_format: bool = False):
        _first_row_no, _last_row_no, _first_col_no, _last_col_no = 0, 0, 0, 0
        if ':' in _range:
            _first_cell_coord = str(_range.split(':')[0])
            _last_cell_coord = str(_range.split(':')[1])
            _first_row_no = int(_first_cell_coord[1:])
            _first_col_no = int(self._sheet_columns.index(_first_cell_coord[0])) + 1
            _last_row_no = int(_last_cell_coord[1:]) + 1
            _last_col_no = int(self._sheet_columns.index(_last_cell_coord[0])) + 2
            self._sheet.merge_cells(_range)
        else:
            _first_row_no = int(_range[1:])
            _first_col_no = int(self._sheet_columns.index(_range[0])) + 1
            _last_row_no = _first_row_no + 1
            _last_col_no = _first_col_no + 1
        self._sheet.cell(row=_first_row_no, column=_first_col_no).value = _value
        self._sheet.cell(row=_first_row_no, column=_first_col_no).alignment = self._align_center
        if _number_format:
            self._sheet.cell(row=_first_row_no, column=_first_col_no).number_format = '#,##0'
        if _fill:
            self._sheet.cell(row=_first_row_no, column=_first_col_no).fill = self._fill_gray
        self._sheet.cell(row=_first_row_no, column=_first_col_no).font = _font_size
        for _row in range(_first_row_no, _last_row_no):
            for _col in range(_first_col_no, _last_col_no):
                self._sheet.cell(row=_row, column=_col).border = self._board

    def run(self):
        try:
            if os.path.isfile(self._temp_path):
                os.remove(self._temp_path)
            _print_timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self._excel = openpyxl.Workbook()
            self._excel.active.title = 'Sheet'
            self._sheet = self._excel.active
            self._sheet.page_setup.paperSize = self._sheet.PAPERSIZE_A4
            self._sheet.page_setup.fitToHeight = 0
            self._sheet.page_setup.fitToWidth = 1
            self._sheet.page_margins.left = 0.2
            self._sheet.page_margins.right = 0.2
            self._sheet.page_margins.top = 0.3
            self._sheet.page_margins.bottom = 1
            self._sheet.print_options.horizontalCentered = True
            self._sheet.oddFooter.left.text = f'출력날짜: {_print_timestamp}\n조회날짜: {self._head_data[0]}'
            self._sheet.oddFooter.right.text = f'{self._head_data[1]} ({len(self._body_data)})\n&[Page] / &N 장'
            for _col in self._sheet_columns:
                if _col == 'A':
                    self._sheet.column_dimensions[_col].width = 4
                elif _col == 'B':
                    self._sheet.column_dimensions[_col].width = 21.9
                else:
                    self._sheet.column_dimensions[_col].width = 7.3
            self.design_cell('B1:B2', '출력날짜', self._font_12, True)
            self.design_cell('C1:E2', '조회날짜', self._font_12, True)
            self.design_cell('F1:H2', 'B/L 번호', self._font_12, True)
            self.design_cell('I1:I2', '면장 개수', self._font_9, True)
            self.design_cell('J1:J2', '선적 대수', self._font_9, True)
            self.design_cell('K1:K2', '선적 중량', self._font_9, True)
            self.design_cell('B3:B4', _print_timestamp, self._font_12)
            self.design_cell('C3:E4', self._head_data[0], self._font_12)
            self.design_cell('F3:H4', self._head_data[1], self._font_12)
            self.design_cell('I3:I4', len(self._body_data), self._font_9)
            self.design_cell('A6:A8', '번호', self._font_7, True)
            self.design_cell('B6:E6', '통관사항', self._font_7, True)
            self.design_cell('F6:K6', '선적사항', self._font_7, True)
            self.design_cell('B7', '수출자', self._font_7, True)
            self.design_cell('C7:D7', '수리일자', self._font_7, True)
            self.design_cell('E7', '통관포장개수', self._font_7, True)
            self.design_cell('B8', '수출신고번호', self._font_7, True)
            self.design_cell('C8:D8', '적재의무기한', self._font_7, True)
            self.design_cell('E8', '통관중량', self._font_7, True)
            self.design_cell('F7:G8', '적하목록관리번호', self._font_7, True)
            self.design_cell('H7:I7', '선기적지', self._font_7, True)
            self.design_cell('J7', '선기적포장개수', self._font_6, True)
            self.design_cell('K7', '분할회수', self._font_7, True)
            self.design_cell('H8:I8', '출항일자', self._font_7, True)
            self.design_cell('J8', '선기적중량', self._font_7, True)
            self.design_cell('K8', '선기적완료여부', self._font_6, True)
            _row_no = 9
            _pre_amount = 0
            _pre_weight = 0
            for i, _row in enumerate(self._body_data):
                self.design_cell(f'A{_row_no}:A{_row_no + 1}', i + 1, self._font_12)
                self.design_cell(f'B{_row_no}', _row[0], self._font_12)
                self.design_cell(f'C{_row_no}:D{_row_no}', _row[1], self._font_12)
                self.design_cell(f'E{_row_no}', _row[2], self._font_12, _number_format=True)
                self.design_cell(f'B{(_row_no + 1)}', _row[6], self._font_12)
                self.design_cell(f'C{_row_no + 1}:D{_row_no + 1}', _row[7], self._font_12)
                self.design_cell(f'E{_row_no + 1}', _row[8], self._font_12, _number_format=True)
                self.design_cell(f'F{_row_no}:G{_row_no + 1}', _row[12], self._font_12)
                self.design_cell(f'H{_row_no}:I{_row_no}', _row[3], self._font_12)
                self.design_cell(f'J{_row_no}', _row[4], self._font_12, _number_format=True)
                self.design_cell(f'K{_row_no}', _row[5], self._font_12)
                self.design_cell(f'H{_row_no + 1}:I{_row_no + 1}', _row[9], self._font_12)
                self.design_cell(f'J{_row_no + 1}', _row[10], self._font_12, _number_format=True)
                self.design_cell(f'K{_row_no + 1}', _row[11], self._font_12)
                _pre_amount += int(_row[4])
                _pre_weight += int(_row[10])
                _row_no += 2
            self.design_cell('J3:J4', _pre_amount, self._font_9, _number_format=True)
            self.design_cell('K3:K4', _pre_weight, self._font_9, _number_format=True)
            self._excel.save(self._temp_path)
            win32api.ShellExecute(
                0, 'print', self._temp_path, f'/d:"{self._printer_name}"', '.', 0
            )
            self.send(False, True, None)
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [self._head_data, self._body_data]
            })
        finally:
            if self._excel is not None:
                self._excel.close()


class DownloadBlDataToPrintForm(_ThreadModel):
    def __init__(self, _thread_queue: Queue, _download_path: str, _excel_name: str, _lookup_index: int, _app_queue: [Queue, None] = None):
        self._code, self._download_path, self._excel_name, self._lookup_index = 'ModelCommon_DownloadBlDataToForm', _download_path, _excel_name, _lookup_index
        self._dbm = _DataBaseManager(self._code)
        super().__init__(_thread_queue, _app_queue)
        self._excel: [openpyxl.Workbook, None] = None
        self._sheet = None
        self._sheet_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        self._font_6 = Font(size=6)
        self._font_7 = Font(size=7)
        self._font_9 = Font(size=9)
        self._font_12 = Font(size=12)
        self._fill_gray = PatternFill(fgColor='F2F2F2', fill_type='solid')
        self._align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self._board = Border(
            left=Side(border_style='thin', color='00000000'),
            right=Side(border_style='thin', color='00000000'),
            top=Side(border_style='thin', color='00000000'),
            bottom=Side(border_style='thin', color='00000000')
        )

    def design_cell(self, _range: str, _value: [str, int], _font_size: Font, _fill: bool = False, _number_format: bool = False, _date_format: bool = False):
        _first_row_no, _last_row_no, _first_col_no, _last_col_no = 0, 0, 0, 0
        if ':' in _range:
            _first_cell_coord = str(_range.split(':')[0])
            _last_cell_coord = str(_range.split(':')[1])
            _first_row_no = int(_first_cell_coord[1:])
            _first_col_no = int(self._sheet_columns.index(_first_cell_coord[0])) + 1
            _last_row_no = int(_last_cell_coord[1:]) + 1
            _last_col_no = int(self._sheet_columns.index(_last_cell_coord[0])) + 2
            self._sheet.merge_cells(_range)
        else:
            _first_row_no = int(_range[1:])
            _first_col_no = int(self._sheet_columns.index(_range[0])) + 1
            _last_row_no = _first_row_no + 1
            _last_col_no = _first_col_no + 1
        self._sheet.cell(row=_first_row_no, column=_first_col_no).value = _value
        self._sheet.cell(row=_first_row_no, column=_first_col_no).alignment = self._align_center
        if _number_format:
            self._sheet.cell(row=_first_row_no, column=_first_col_no).number_format = '#,##0'
        if _date_format:
            self._sheet.cell(row=_first_row_no, column=_first_col_no).number_format = 'yyyy-mm-dd hh:mm:ss'
        if _fill:
            self._sheet.cell(row=_first_row_no, column=_first_col_no).fill = self._fill_gray
        self._sheet.cell(row=_first_row_no, column=_first_col_no).font = _font_size
        for _row in range(_first_row_no, _last_row_no):
            for _col in range(_first_col_no, _last_col_no):
                self._sheet.cell(row=_row, column=_col).border = self._board

    def run(self):
        try:
            self._excel = openpyxl.Workbook()
            self._excel.remove_sheet(self._excel.active)
            _dcr = self._dbm.connecting()
            if _dcr[0]:
                _query = f'''
                SELECT EXCEL_ROW_NO,
                       lb.LOOKUP_DATE,
                       BL_NUMBER,
                       COUNT(*),
                       SUM(PRE_AMOUNT),
                       SUM(PRE_WEIGHT)
                FROM merged_ei_bn JOIN lookup_bl lb ON merged_ei_bn.LOOKUP_BL_INDEX = lb.l_bl_index
                WHERE LOOKUP_BL_INDEX = {self._lookup_index}
                GROUP BY BL_NUMBER, EXCEL_ROW_NO
                ORDER BY EXCEL_ROW_NO;
                '''
                _der = self._dbm.execute_query(_query)
                if _der[0]:
                    for _sheet_info in _der[1]:
                        self._excel.create_sheet(_sheet_info[2])
                        self._sheet = self._excel[_sheet_info[2]]
                        self._sheet.page_setup.paperSize = self._sheet.PAPERSIZE_A4
                        self._sheet.page_setup.fitToHeight = 0
                        self._sheet.page_setup.fitToWidth = 1
                        self._sheet.page_margins.left = 0.2
                        self._sheet.page_margins.right = 0.2
                        self._sheet.page_margins.top = 0.3
                        self._sheet.page_margins.bottom = 1
                        self._sheet.print_options.horizontalCentered = True
                        self._sheet.oddFooter.left.text = f'출력날짜: =NOW()\n조회날짜: {_sheet_info[1]}'
                        self._sheet.oddFooter.right.text = f'{_sheet_info[2]} ({_sheet_info[3]})\n&[Page] / &N 장'
                        for _col in self._sheet_columns:
                            if _col == 'A':
                                self._sheet.column_dimensions[_col].width = 4
                            elif _col == 'B':
                                self._sheet.column_dimensions[_col].width = 21.9
                            else:
                                self._sheet.column_dimensions[_col].width = 7.3
                        self.design_cell('B1:B2', '출력날짜', self._font_12, True)
                        self.design_cell('C1:E2', '조회날짜', self._font_12, True)
                        self.design_cell('F1:H2', 'B/L 번호', self._font_12, True)
                        self.design_cell('I1:I2', '면장 개수', self._font_9, True)
                        self.design_cell('J1:J2', '선적 대수', self._font_9, True)
                        self.design_cell('K1:K2', '선적 중량', self._font_9, True)
                        self.design_cell('B3:B4', '=NOW()', self._font_12, _date_format=True)
                        self.design_cell('C3:E4', _sheet_info[1], self._font_12)
                        self.design_cell('F3:H4', _sheet_info[2], self._font_12)
                        self.design_cell('I3:I4', _sheet_info[3], self._font_9)
                        self.design_cell('J3:J4', _sheet_info[4], self._font_9, _number_format=True)
                        self.design_cell('K3:K4', _sheet_info[5], self._font_9, _number_format=True)
                        self.design_cell('A6:A8', '번호', self._font_7, True)
                        self.design_cell('B6:E6', '통관사항', self._font_7, True)
                        self.design_cell('F6:K6', '선적사항', self._font_7, True)
                        self.design_cell('B7', '수출자', self._font_7, True)
                        self.design_cell('C7:D7', '수리일자', self._font_7, True)
                        self.design_cell('E7', '통관포장개수', self._font_7, True)
                        self.design_cell('B8', '수출신고번호', self._font_7, True)
                        self.design_cell('C8:D8', '적재의무기한', self._font_7, True)
                        self.design_cell('E8', '통관중량', self._font_7, True)
                        self.design_cell('F7:G8', '적하목록관리번호', self._font_7, True)
                        self.design_cell('H7:I7', '선기적지', self._font_7, True)
                        self.design_cell('J7', '선기적포장개수', self._font_6, True)
                        self.design_cell('K7', '분할회수', self._font_7, True)
                        self.design_cell('H8:I8', '출항일자', self._font_7, True)
                        self.design_cell('J8', '선기적중량', self._font_7, True)
                        self.design_cell('K8', '선기적완료여부', self._font_6, True)
                        _list_query = f'''
                        SELECT EXPORT_SHIPPER,
                               ACCEPT_DATE,
                               TOTAL_AMOUNT,
                               SHIPMENT_PLACE,
                               PRE_AMOUNT,
                               DIVISION_COLLECT,
                               FORMAT_EXPORT_NUMBER,
                               LOAD_DUTY_DEADLINE,
                               TOTAL_WEIGHT,
                               DEPARTURE_DATE,
                               PRE_WEIGHT,
                               FLAG_PRE_SHIPPING,
                               MANAGEMENT_NUMBER
                        FROM merged_ei_bn
                        WHERE BL_NUMBER = '{_sheet_info[2]}'
                          AND LOOKUP_BL_INDEX = {self._lookup_index}
                        '''
                        _l_der = self._dbm.execute_query(_list_query)
                        if _l_der[0]:
                            _row_no = 9
                            for i, _row in enumerate(_l_der[1]):
                                self.design_cell(f'A{_row_no}:A{_row_no + 1}', i + 1, self._font_12)
                                self.design_cell(f'B{_row_no}', _row[0], self._font_12)
                                self.design_cell(f'C{_row_no}:D{_row_no}', _row[1], self._font_12)
                                self.design_cell(f'E{_row_no}', _row[2], self._font_12, _number_format=True)
                                self.design_cell(f'B{(_row_no + 1)}', _row[6], self._font_12)
                                self.design_cell(f'C{_row_no + 1}:D{_row_no + 1}', _row[7], self._font_12)
                                self.design_cell(f'E{_row_no + 1}', _row[8], self._font_12, _number_format=True)
                                self.design_cell(f'F{_row_no}:G{_row_no + 1}', _row[12], self._font_12)
                                self.design_cell(f'H{_row_no}:I{_row_no}', _row[3], self._font_12)
                                self.design_cell(f'J{_row_no}', _row[4], self._font_12, _number_format=True)
                                self.design_cell(f'K{_row_no}', _row[5], self._font_12)
                                self.design_cell(f'H{_row_no + 1}:I{_row_no + 1}', _row[9], self._font_12)
                                self.design_cell(f'J{_row_no + 1}', _row[10], self._font_12, _number_format=True)
                                self.design_cell(f'K{_row_no + 1}', _row[11], self._font_12)
                                _row_no += 2
                        else:
                            self.send(False, False, _der[2])
                            break
                    self._excel.save(f'{self._download_path}/BL_PRINT_{self._excel_name}')
                    self.send(False, True, f'{self._download_path}/BL_PRINT_{self._excel_name}')
                else:
                    self.send(False, False, _der[2])
            else:
                self.send(False, False, _dcr[1])
        except:
            self.send(False, False, {
                'ECD': self._code + '_Run',
                'CLS': sys.exc_info()[0],
                'DES': sys.exc_info()[1],
                'LNO': sys.exc_info()[2].tb_lineno,
                'DTA': [self._download_path, self._excel_name, self._lookup_index]
            })
        finally:
            if self._excel is not None:
                self._excel.close()
            self._dbm.disconnecting()


